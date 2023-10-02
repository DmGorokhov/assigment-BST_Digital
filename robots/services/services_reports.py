from io import BytesIO
from itertools import groupby
from openpyxl import Workbook
from django.db.models import Count
from datetime import timedelta, datetime
from robots.models import Robot


def make_week_report():
    end_date = datetime.now()
    start_date = end_date - timedelta(days=7)

    robots_by_week = Robot.objects.filter(
        created__range=[start_date, end_date]).values('model', 'version'). \
        annotate(count=Count('serial')).order_by('model')

    robots_group_by_model = groupby(robots_by_week, key=lambda x: x['model'])
    workbook = Workbook()

    for model, group in robots_group_by_model:
        sheet = workbook.create_sheet(title=model)
        sheet['A1'] = 'Модель'
        sheet['B1'] = 'Версия'
        sheet['C1'] = 'Количество за неделю'
        row_index = 2
        for item in group:
            sheet.cell(row=row_index, column=1, value=item['model'])
            sheet.cell(row=row_index, column=2, value=item['version'])
            sheet.cell(row=row_index, column=3, value=item['count'])
            row_index += 1
    report_file = BytesIO()
    workbook.save(report_file)
    report_file.seek(0)
    return report_file
