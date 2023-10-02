from io import BytesIO
from itertools import groupby
from openpyxl import Workbook
from django.db.models import Count
from datetime import timedelta
import json
from pydantic import BaseModel, constr, ValidationError
from datetime import datetime
from .models import Robot
from django.db import DatabaseError
from typing import Literal


class BaseRobotData(BaseModel):
    model: Literal['R1', 'R2', 'X5', '13', 'DD', 'S', '07']
    version: constr(max_length=2)
    created: datetime


class NewRobot(BaseRobotData):
    serial: constr(max_length=5)


class ParsedRobot(BaseModel):
    clean_data: dict = None
    error: list = None


class NewRobotResponse(BaseModel):
    response_message: dict
    response_status: int


def get_parsed_and_validate_robot(new_robot_data: json) -> ParsedRobot:
    try:
        robot_request = BaseRobotData.model_validate_json(new_robot_data)
        return ParsedRobot(clean_data=robot_request.model_dump())
    except ValidationError as er:
        return ParsedRobot(error=er.errors())


def save_new_robot(robot_data: json) -> NewRobotResponse:
    try:
        robot_serial = f'{robot_data.get("model")}-{robot_data.get("version")}'
        new_robot_data = NewRobot(**robot_data, serial=robot_serial)
        Robot.objects.create(**new_robot_data.model_dump())
        return NewRobotResponse(response_message={'success': 'Robot added'},
                                response_status=200)
    except DatabaseError:
        return NewRobotResponse(response_message={
            'error_message': 'Internal Server Error'},
            response_status=500)


def make_week_report():
    end_date = datetime.now()
    start_date = end_date - timedelta(days=7)

    robots_by_week = Robot.objects.filter(
        created__range=[start_date, end_date]).values('model', 'version'). \
        annotate(count=Count('serial')).order_by('model')

    robots_group_by_model = groupby(robots_by_week, key=lambda x: x['model'])
    workbook = Workbook()

    for model, group in robots_group_by_model:
        sheet = workbook.create_sheet(title=model)
        sheet['A1'] = 'Модель'
        sheet['B1'] = 'Версия'
        sheet['C1'] = 'Количество за неделю'
        row_index = 2
        for item in group:
            sheet.cell(row=row_index, column=1, value=item['model'])
            sheet.cell(row=row_index, column=2, value=item['version'])
            sheet.cell(row=row_index, column=3, value=item['count'])
            row_index += 1
    report_file = BytesIO()
    workbook.save(report_file)
    report_file.seek(0)
    return report_file
